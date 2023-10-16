import openpyxl
import configparser
import locale
import time
# import pyfiglet

# def create_banner(text, font="big"):
#     banner = pyfiglet.figlet_format(text, font=font)
#     return banner

# Set the locale for currency formatting (Indonesian Rupiah)
locale.setlocale(locale.LC_ALL, 'id_ID')

# Load exchange rates from the config file
config = configparser.ConfigParser()
config.read('config.ini')

# banner
# program_name = "Exchange Rate"
# banner = create_banner(program_name)
# print(banner)

# usd_to_idr = float(config['exchange_rates']['USD_TO_IDR'])
usd_to_idr = float(input("Enter exchange rates USD to IDR: "))

# input_file = config['setup']['input']
input_file = input("Enter input file: ")
if ".xlsx" not in input_file:
    input_file = input_file + ".xlsx"

# output_file = config['setup']['output']
output_file = input("Enter output file: ")
if ".xlsx" not in output_file:
    output_file = output_file + ".xlsx"

print("Please wait...")
time.sleep(1)


# Open the input Excel file
usd_file_path = input_file 
usd_wb = openpyxl.load_workbook(usd_file_path)
usd_sheet = usd_wb.active

# Create a new output Excel file
idr_wb = openpyxl.Workbook()
idr_sheet = idr_wb.active

# Copy the header row from the input to the output
for row in usd_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    idr_sheet.append(row)

# Specify the column to convert (e.g., column D)
column_to_convert = 'D'

# Iterate through rows in the input Excel file and perform the conversion
for row in usd_sheet.iter_rows(min_row=2, values_only=True):
    idr_row = list(row)
    usd_gaji = idr_row[usd_sheet[column_to_convert][0].column - 1]

    if usd_gaji is not None:
        idr_gaji = usd_gaji * usd_to_idr
        # Format the IDR salary as Rupiah
        idr_gaji_formatted = locale.currency(idr_gaji, grouping=True)
        idr_row[usd_sheet[column_to_convert][0].column - 1] = idr_gaji_formatted
    idr_sheet.append(idr_row)

# Save the new output workbook with the same filename as the input
output_file_path = output_file
idr_wb.save(output_file_path)

# Close the workbooks
usd_wb.close()
idr_wb.close()

